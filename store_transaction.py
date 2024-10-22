from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import MainIO as io
from MainIO import date
def add_recurring(name=None, date=None, ie=None, value=None, category=None):
    workbook = load_workbook('NSB-AlgoProg-Group-Work/Recurring.xlsx')
    workingfile = workbook.active
    
    # ADD NEW DATA FROM ABOVE TO NEW LINE IN EXCEL
    row = 2
    col = 2
    while workingfile.cell(row=row, column=col).value != None:
        row += 1

    workingfile[get_column_letter(col) + str(row)]= name.upper()
    workingfile[get_column_letter(col+1) + str(row)] = io.date()
    workingfile[get_column_letter(col+2) + str(row)] = ie.upper()
    workingfile[get_column_letter(col+3) + str(row)] = value
    workingfile[get_column_letter(col+4) + str(row)] = category.upper()

    workbook.save('NSB-AlgoProg-Group-Work/Recurring.xlsx')