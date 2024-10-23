import tkinter as tk
from tkinter import ttk

root = tk.Tk()
root.title("Entry Book")
text_var = tk.StringVar()
text_var.set("Digital Entry Book")

label = tk.Label(root, 
                 textvariable=text_var, 
                 anchor=tk.CENTER,      
                 height=3,              
                 width=25,                                
                 font=("Arial", 12, "bold"), 
                 fg="black",                            
                 justify=tk.CENTER,    
                 relief=tk.RAISED,                    
                )
label.pack(pady=10, padx=10)
search_frame = ttk.Frame(root)
search_frame.pack(pady=10)
search_var = tk.StringVar()
search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
search_entry.grid(row=0, column=0, padx=10)
category_var = tk.StringVar()
category_combobox = ttk.Combobox(search_frame, textvariable=category_var, state='readonly', width=20)
category_combobox['values'] = ("Name", "Date", "Income/Expense", "Value", "Category", "Delivered", "ID")
category_combobox.grid(row=0, column=1)
category_combobox.current(0)

def search_data():
    search_term = search_var.get().lower()
    category = category_combobox.get()
    col_map = {
        "Name": 0,
        "Date": 1,
        "Income/Expense": 2,
        "Value": 3,
        "Category": 4,
        "Delivered": 5,
        "ID": 6
    }
    selected_col = col_map[category]
    for item in treeview.get_children():
        treeview.delete(item)
    from openpyxl import load_workbook
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active

    for index, row in enumerate(db.iter_rows(min_row=2, values_only=True), start=1):
        value = str(row[selected_col]).lower() if row[selected_col] is not None else ''
        if search_term in value:
            treeview.insert("", "end", text=str(index), values=(
                row[1] if row[1] is not None else '',
                row[2] if row[2] is not None else '',
                row[3] if row[3] is not None else '',
                row[4] if row[4] is not None else '',
                row[5] if row[5] is not None else '',
                row[6] if row[6] is not None else '',
                row[7] if row[7] is not None else ''
            ))

search_button = ttk.Button(search_frame, text="Search", command=search_data)
search_button.grid(row=0, column=2, padx=10)

treeview = ttk.Treeview(columns=("name", "date", "ie", "value", "category", "delivered", "id"))
treeview.pack(pady=70, padx=30)
treeview.heading("#0", text="Index")
treeview.column("#0", width=50)
treeview.heading("name", text="Name")
treeview.heading("date", text="Date")
treeview.column("date", width=150)
treeview.heading("ie", text="Income/Expense")
treeview.column("ie", width=120)
treeview.heading("value", text="Value")
treeview.heading("category", text="Category")
treeview.heading("delivered", text="Delivered")
treeview.column("delivered", width=100)
treeview.heading("id", text="ID")
treeview.column("id", width=150)

def load_data():
    from openpyxl import load_workbook
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active

    for index, row in enumerate(db.iter_rows(min_row=2, values_only=True), start=1):
        treeview.insert("", "end", text=str(index), values=(
            row[1] if row[1] is not None else '',
            row[2] if row[2] is not None else '',
            row[3] if row[3] is not None else '',
            row[4] if row[4] is not None else '',
            row[5] if row[5] is not None else '',
            row[6] if row[6] is not None else '',
            row[7] if row[7] is not None else ''
        ))

load_data()

label.pack(pady=20, padx=20)
treeview.pack()
root.mainloop()