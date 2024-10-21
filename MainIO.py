"""
    REMINDER:
    - DO NOT FORGET TO INSTALL OPENPYXL IN YOUR MACHINE!
    - INSTALL IN CMD WITH: 'pip install openpyxl' / 'pip3 install openpyxl' /
            'python -m pip install openpyxl' / 'python3 -m pip install openpyxl'.
    - THE MODULE'S FILE MANIPULATOR CAN ONLY WORK WITH .xlsx FILE FORMATS (2010 ABOVE)
      AND CANNOT WORK WITH OTHER EXCEL FILE FORMATS.
    - CODE CANNOT WORK WHEN EXCEL FILE IS BEING OPENED.
    - NEVER EDIT THE VALUES WRITTEN ON 'OTHER INFO'.
"""

# GENERATE A SPECIAL ID KEY TO ALLOW DIFFERENCE BETWEEN ENTRIES

def gen_id_key():
    with open('NSB-AlgoProg-Group-Work/Set_Const.txt', 'r') as file:
        current_value = int(file.readline().strip().split('ID ')[-1])
        new_value = current_value + 1
        id_key = f"{current_value:06d}"
    
    with open('NSB-AlgoProg-Group-Work/Set_Const.txt', 'w') as file:
        file.seek(0)
        file.write(f"ID {new_value:06d}")
        
    return id_key

# GENERATE CURRENT DATE AND TIME

import datetime as dt

def date():
    current_date = dt.datetime.now().date()
    return current_date.strftime("%Y-%m-%d")

def time():
    timezone = dt.timezone(dt.timedelta(hours=7))
    current_time = dt.datetime.now(timezone).time()
    return current_time.strftime("%H:%M:%S")

# ADD INCOME AND EXPENSE ENTRIES TO THE DATABASE

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def balance(new_value, ie, delivered):
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active
    
    balance = 0
    
    for column in db.iter_cols(min_row=2, max_row=2, min_col=12, max_col=12):
        for cell in column:
            if cell.value != None:
                balance = cell.value
    print(f"Current balance: {balance}")
    
    if balance is None:
        balance = 0
        
    if delivered == True:
        if ie == 'I':
            balance += new_value
        elif ie == 'E':
            balance -= new_value

    print(f"New balance: {balance}")
    db[get_column_letter(12) + str(2)].value = balance
    wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')
    wb.close()
    
def income(name, value=0, category='income', delivered=False):
	wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
	db = wb.active 
 
	row = 2
	col = 2
	while db.cell(row=row, column=col).value != None:
		row += 1

	db[get_column_letter(col) + str(row)] = name.upper()
	db[get_column_letter(col+1) + str(row)] = date()
	db[get_column_letter(col+2) + str(row)] = 'I'
	db[get_column_letter(col+3) + str(row)] = value
	db[get_column_letter(col+4) + str(row)] = category.upper()
	db[get_column_letter(col+5) + str(row)] = delivered
	db[get_column_letter(col+6) + str(row)] = gen_id_key()
	db[get_column_letter(col+7) + str(row)] = f"{date()} / {time()}"

	balance(value, 'I', delivered)

	wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')

def expense(name, value=0, category='expense', delivered=False):
	wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
	db = wb.active 
 
	row = 2
	col = 2
	while db.cell(row=row, column=col).value != None:
		row += 1

	db[get_column_letter(col) + str(row)] = name.upper()
	db[get_column_letter(col+1) + str(row)] = date()
	db[get_column_letter(col+2) + str(row)] = 'E'
	db[get_column_letter(col+3) + str(row)] = value
	db[get_column_letter(col+4) + str(row)] = category.upper()
	db[get_column_letter(col+5) + str(row)] = delivered
	db[get_column_letter(col+6) + str(row)] = gen_id_key()
	db[get_column_letter(col+7) + str(row)] = f"{date()} / {time()}"
 
	balance(value, 'E', delivered)
    
	wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')

# MODIFY ENTRIES IN THE DATABASE

def read_all():
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active
    
    for row in db.iter_rows(max_col=8, min_row=2, max_row=db.max_row):
        for cell in row:
            if cell.value != None:
                print(cell.value, end=" ")
        print('', end="\n")

def change(name=None, ie=None, value=None, category=None, delivered=None):
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active
    
    data_list = []
    id_list = []
    date_list = []
    entry_list = []
    total = 0
    
    """print("Available entries:")"""  
    for columnentry in db.iter_cols(min_row=2, min_col=2, max_col=2):
    # GET THE ENTRY NAMES
        for cellentry in columnentry:
            if cellentry.value != None:
                entry_list.append(cellentry)
                total += 1
    
    for columndate in db.iter_cols(min_row=2, min_col=3, max_col=3):
    # GET THE ENTRY DATES
        for celldate in columndate:
            if celldate.value != None:
                date_list.append(celldate)
                total += 1
                
    for columnid in db.iter_cols(min_row=2, min_col=8, max_col=8):
    # GET ID OF ENTRIES
        for cellid in columnid:
            if cellid.value != None:
                id_list.append(cellid)
                total += 1
    
    total_entry = total // 3
    
    for each_entry in range(total_entry):
        inner_list = []
        inner_list.append(entry_list[each_entry])
        inner_list.append(date_list[each_entry])
        inner_list.append(id_list[each_entry])
        data_list.append(inner_list)
        print(id_list[each_entry], date_list[each_entry], entry_list[each_entry], end='\n')
    
    entryselect = str(input('Select entry: ')).upper().strip()
    dateselect = str(input('Select date (type by exact format): ')).strip()
    idselect = str(input('Select ID (type by exact format): ')).strip() 

    for every_entry in data_list:
        if every_entry[2] == entryselect:
            if every_entry[1] == dateselect:
                if every_entry[0] == idselect:
                    print('Entry found.')
                    index = data_list.index(every_entry) + 2
                    db[get_column_letter(2) + str(index)] = name if name != None else db[get_column_letter(2) + str(index)].value
                    db[get_column_letter(3) + str(index)] = ie.upper() if ie != None else db[get_column_letter(3) + str(index)].value
                    db[get_column_letter(4) + str(index)] = value if value != None else db[get_column_letter(4) + str(index)].value
                    db[get_column_letter(5) + str(index)] = category.upper() if category != None else db[get_column_letter(5) + str(index)].value
                    db[get_column_letter(6) + str(index)] = delivered if delivered != None else db[get_column_letter(6) + str(index)].value
                    db[get_column_letter(9) + str(index)] = f"{date()} / {time()}"
                    balance(value, ie, delivered)
                    break
                else:
                    print('ID does not match.')
            else:
                print('Date does not match.')
        else:
            print('Entry does not match.')
        
    read_all()

    wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')
    
def delete():
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active

    entry_list = []
    
    """print("Available entries:")"""
    for column in db.iter_cols(min_row=2, max_row=db.max_row, min_col=2, max_col=2):
        for cell in column:
            if cell.value != None:
                entry_list.append(cell.value)

    print(entry_list)

    selected = str(input('Select entry to delete: ')).strip().upper()

    if selected not in entry_list:
        print('Entry does not exist.')
    else:
        index = entry_list.index(selected) + 2 
        db[get_column_letter(2) + str(index)].value = None
        db[get_column_letter(3) + str(index)].value = None
        db[get_column_letter(4) + str(index)].value = None
        db[get_column_letter(5) + str(index)].value = None
        db[get_column_letter(6) + str(index)].value = None
        db[get_column_letter(7) + str(index)].value = None
        db[get_column_letter(8) + str(index)].value = None
        balance(db[get_column_letter(5) + str(index)].value, 'E' if db[get_column_letter(3) + str(index)].value == 'I' else 'I', db[get_column_letter(7) + str(index)].value)
        
    read_all()
    
    wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')

# SEARCH ENTRIES IN THE DATABASE

def search(file_path='Database.xlsx'):
    wb = load_workbook(f'NSB-AlgoProg-Group-Work/{file_path}')
    db = wb.active
    
    data_list = []
    name_list = []
    date_list = []
    ie_list = []
    value_list = []
    category_list = []
    delivered_list = []
    id_list = []
    total = 0
    
    """print("Available entries:")"""  
    for columnname in db.iter_cols(min_row=2, min_col=2, max_col=2):
    # GET THE ENTRY NAMES
        for cellname in columnname:
            if cellname.value != None:
                name_list.append(cellname.value)
                total += 1
    
    for columndate in db.iter_cols(min_row=2, min_col=3, max_col=3):
    # GET THE ENTRY DATES
        for celldate in columndate:
            if celldate.value != None:
                date_list.append(celldate.value)
                total += 1
    
    for columnie in db.iter_cols(min_row=2, min_col=4, max_col=4):
    # GET THE ENTRY TYPE
        for cellie in columnie:
            if cellie.value != None:
                ie_list.append(cellie.value)
                total += 1

    for columnvalue in db.iter_cols(min_row=2, min_col=5, max_col=5):
    # GET THE ENTRY VALUES
        for cellvalue in columnvalue:
            if cellvalue.value != None:
                value_list.append(cellvalue.value)
                total += 1
    
    for columncategory in db.iter_cols(min_row=2, min_col=6, max_col=6):
    # GET THE ENTRY CATEGORIES
        for cellcategory in columncategory:
            if cellcategory.value != None:
                category_list.append(cellcategory.value)
                total += 1
    
    if file_path == 'Database.xlsx':
        for columndelivered in db.iter_cols(min_row=2, min_col=7, max_col=7):
        # GET THE ENTRY DELIVERY STATUS
            for celldelivered in columndelivered:
                if celldelivered.value != None:
                    delivered_list.append(celldelivered.value)
                    total += 1
        
        for columnid in db.iter_cols(min_row=2, min_col=8, max_col=8):
        # GET ID OF ENTRIES
            for cellid in columnid:
                if cellid.value != None:
                    id_list.append(cellid.value)
                    total += 1
    
    total_entry = total // 7 if file_path == 'Database.xlsx' else total // 5
    
    for each_entry in range(total_entry):
        inner_list = []
        inner_list.append(name_list[each_entry])
        inner_list.append(date_list[each_entry])
        inner_list.append(ie_list[each_entry])
        inner_list.append(value_list[each_entry])
        inner_list.append(category_list[each_entry])
        inner_list.append(delivered_list[each_entry]) if file_path == 'Database.xlsx' else next
        inner_list.append(id_list[each_entry]) if file_path == 'Database.xlsx' else next
        data_list.append(inner_list)
    
    read_all()
    
    typeofentry = str(input('Type of entry: ')).upper().strip()
    entryselect = str(input('Select entry: ')).upper().strip() if (typeofentry != 'VALUE') else str(input('Select value: ').strip()).split()
    
    for every_entry in data_list:
        match typeofentry:
            case 'NAME':
                if every_entry[0] == entryselect:
                    # OUTPUT THE ENTRY
                    print(every_entry)
            case 'DATE':
                if every_entry[1] == entryselect:
                    # OUTPUT THE ENTRY
                    print(every_entry)
            case 'IE':
                if every_entry[2] == entryselect:
                    # OUTPUT THE ENTRY
                    print(every_entry)
            case 'VALUE':
                # FORMAT: "SYMBOL AMOUNT"
                match entryselect[0]:
                    case '>':
                        # OUTPUT THE ENTRY
                        print(every_entry) if every_entry[3] > float(entryselect[1]) else next
                    case '<':
                        # OUTPUT THE ENTRY
                        print(every_entry) if every_entry[3] < float(entryselect[1]) else next
                    case '>=':
                        # OUTPUT THE ENTRY
                        print(every_entry) if every_entry[3] >= float(entryselect[1]) else next
                    case '<=':
                        # OUTPUT THE ENTRY
                        print(every_entry) if every_entry[3] <= float(entryselect[1]) else next
                    case '==':
                        # OUTPUT THE ENTRY
                        print(every_entry) if every_entry[3] == float(entryselect[1]) else next
            case 'CATEGORY':
                if every_entry[4] == entryselect:
                    # OUTPUT THE ENTRY
                    print(every_entry)
            case 'DELIVERED':
                if every_entry[5] == entryselect:
                    # OUTPUT THE ENTRY
                    print(every_entry)
            case 'ID':
                if every_entry[6] == entryselect:
                    # OUTPUT THE ENTRY
                    print(every_entry)
            case other:
                print('Invalid type of entry.')

    wb.close()

# TESTING
search()

# NOTE:
# CODE HAS NOT BEEN FULLY TESTED YET // STILL NEED MORE TESTING
# ADDITION OF BALANCE FUNCTIONALITY IS REQUIRED