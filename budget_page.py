import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl as ox
import budget as bdgt
class BudgetPage:
    def __init__(self, master):
        self.master = master
        self.master.title("Budget Page")
        
        treeview = ttk.Treeview()
        treeview = ttk.Treeview(columns=("name", "date", "ie", "value", "category", "delivered", "id"))
        treeview.pack(pady=70, padx=30)
        treeview.heading("#0", text="Index")
        treeview.column("#0", width=50)
        treeview.heading("name", text="Name")
        treeview.heading("date", text="Date")
        treeview.column("date", width=150)
        treeview.heading("ie", text="Income/Expense")
        treeview.column("ie", width=120)
        treeview.heading("value", text="Value")
        treeview.heading("category", text="Category")
        treeview.heading("delivered", text="Delivered")
        treeview.column("delivered", width=100)
        treeview.heading("id", text="ID")
        treeview.column("id", width=150)

        from openpyxl import load_workbook
        wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
        db = wb.active
        
        for index, row in enumerate(db.iter_rows(min_row=2, values_only=True), start=1):
            if row[3] == 'E':
                treeview.insert("", "end", text=str(index), values=(
                    row[1] if row[1] is not None else '',
                    row[2] if row[2] is not None else '',
                    row[3] if row[3] is not None else '',
                    row[4] if row[4] is not None else '',
                    row[5] if row[5] is not None else '',
                    row[6] if row[6] is not None else '',
                    row[7] if row[7] is not None else ''
                ))
        for item in treeview.get_children():
            print(treeview.item(item)['values'])
        
        self.initial_budget = bdgt.get_limit()
        self.initial_budget_label = tk.Label(self.master, text=f"Initial Budget: ${self.initial_budget}", font=("Arial", 16, "bold"))
        self.initial_budget_label.place(x=50, y=310)
        self.period_spending = bdgt.get_spending()
        self.period_spending_label = tk.Label(self.master, text=f"Total Spending: ${self.period_spending}", font=("Arial", 16, "bold"))
        self.period_spending_label.place(x=50, y=340)
        self.remaining = self.initial_budget - self.period_spending
        self.remaining_label = tk.Label(self.master, text=f"Remaining Budget: ${self.remaining}", font=("Arial", 16, "bold"))
        self.remaining_label.place(x=50, y=370)

if __name__ == "__main__":
    root = tk.Tk()
    app = BudgetPage(root)
    root.mainloop()