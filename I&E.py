"""
    REMINDER:
    - DO NOT FORGET TO INSTALL OPENPYXL IN YOUR MACHINE!
    - INSTALL IN CMD WITH: 'pip install openpyxl' / 'pip3 install openpyxl' /
            'python -m pip install openpyxl' / 'python3 -m pip install openpyxl'.
    - THE MODULE'S FILE MANIPULATOR CAN ONLY WORK WITH .xlsx FILE FORMATS (2010 ABOVE)
      AND CANNOT WORK WITH OTHER EXCEL FILE FORMATS.
    - CODE CANNOT WORK WHEN EXCEL FILE IS BEING OPENED.
    - NEVER EDIT THE VALUES WRITTEN ON 'OTHER INFO'.
"""

# GENERATE A SPECIAL ID KEY TO ALLOW DIFFERENCE BETWEEN ENTRIES

def gen_id_key():
    with open('NSB-AlgoProg-Group-Work/Set_Const.txt', 'r') as file:
        current_value = int(file.readline().strip().split('ID ')[-1])
        new_value = current_value + 1
        id_key = f"{current_value:06d}"
    
    with open('NSB-AlgoProg-Group-Work/Set_Const.txt', 'w') as file:
        file.seek(0)
        file.write(f"ID {new_value:06d}")
        
    return id_key

# GENERATE CURRENT DATE AND TIME

import datetime as dt

def date():
    current_date = dt.datetime.now().date()
    return current_date.strftime("%Y-%m-%d")

def time():
    timezone = dt.timezone(dt.timedelta(hours=7))
    current_time = dt.datetime.now(timezone).time()
    return current_time.strftime("%H:%M:%S")

# ADD INCOME AND EXPENSE ENTRIES TO THE DATABASE

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def balance(new_value, ie, delivered):
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active
    
    balance = 0
    
    for column in db.iter_cols(min_row=2, max_row=2, min_col=12, max_col=12):
        for cell in column:
            if cell.value != None:
                balance = cell.value
    print(f"Current balance: {balance}")
    
    if balance is None:
        balance = 0
        
    if delivered == True:
        if ie == 'I':
            balance += new_value
        elif ie == 'E':
            balance -= new_value

    print(f"New balance: {balance}")
    db[get_column_letter(12) + str(2)].value = balance
    wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')
    wb.close()
    
def income(name, value=0, category='income', delivered=False):
	wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
	db = wb.active 
 
	row = 2
	col = 2
	while db.cell(row=row, column=col).value != None:
		row += 1

	db[get_column_letter(col) + str(row)] = name.upper()
	db[get_column_letter(col+1) + str(row)] = date()
	db[get_column_letter(col+2) + str(row)] = 'I'
	db[get_column_letter(col+3) + str(row)] = value
	db[get_column_letter(col+4) + str(row)] = category.upper()
	db[get_column_letter(col+5) + str(row)] = delivered
	db[get_column_letter(col+6) + str(row)] = gen_id_key()
	db[get_column_letter(col+7) + str(row)] = f"{date()} / {time()}"

	balance(value, 'I', delivered)

	wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')

def expense(name, value=0, category='expense', delivered=False):
	wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
	db = wb.active 
 
	row = 2
	col = 2
	while db.cell(row=row, column=col).value != None:
		row += 1

	db[get_column_letter(col) + str(row)] = name.upper()
	db[get_column_letter(col+1) + str(row)] = date()
	db[get_column_letter(col+2) + str(row)] = 'E'
	db[get_column_letter(col+3) + str(row)] = value
	db[get_column_letter(col+4) + str(row)] = category.upper()
	db[get_column_letter(col+5) + str(row)] = delivered
	db[get_column_letter(col+6) + str(row)] = gen_id_key()
	db[get_column_letter(col+7) + str(row)] = f"{date()} / {time()}"
 
	balance(value, 'E', delivered)
    
	wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')

# MODIFY ENTRIES IN THE DATABASE

def read():
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active
    
    for row in db.iter_rows(max_col=8, min_row=2, max_row=db.max_row):
        for cell in row:
            if cell.value != None:
                print(cell.value, end=" ")
        print('', end="\n")

def change(name=None, ie=None, value=None, category=None, delivered=None):
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active

    entry_list = []
    
    print("Available entries:")  
    for column in db.iter_cols(min_row=2, min_col=3, max_col=3):
    # GET ID OF ENTRIES FIRST
        for cell in column:
            if cell.value != None:
                for columnid in db.iter_cols(min_row=2, min_col=2, max_col=2):
                # THEN GET THE ENTRY NAMES OF THEIR CORRESPONDING IDs
                    for cellid in columnid:
                        entry_list[cell] = cellid
                        print(cell, cellid, end='\n')
    
    #INCOMPLETE // NEED TO FIND ID ACCORDING TO ROW NUMBER POSITION OF ENTRY
    
    entryselect = str(input('Select entry to change: ')).upper().strip()
    dateselect = str(input('Select date to change (type by exact format): ')).strip()
    idselect = str(input('Select ID to change (type by exact format): ')).strip() 

    # NEED TO FIND ROW NUMBER POSITION OF ENTRY BY ID OR DATE
    
    # NEED IF STATEMENT FOR VALIDATING BOTH DATE & ID FORMAT
    
    # NEED TO STORE ENTRY, DATE, & ID IN ONE INNER LIST, PER ENTRY, THEN STORE ALL INNER LISTS IN 'entry_list'
    
    # NEED TO FIND THE CORRESPONDING ENTRY, DATE, & ID IN 'entry_list'
    if entryselect not in entry_list:
        print('Entry does not exist.')
    else:
        if dateselect not in entry_list.get(key=entryselect):
            print('Date with the corresponding entry does not exist.')
        else:
            index = entry_list.index(entryselect)
            db[get_column_letter(2) + str(index)] = name if name != None else db[get_column_letter(2) + str(index)].value
            db[get_column_letter(3) + str(index)] = ie.upper() if ie != None else db[get_column_letter(3) + str(index)].value
            db[get_column_letter(4) + str(index)] = value if value != None else db[get_column_letter(4) + str(index)].value
            db[get_column_letter(5) + str(index)] = category.upper() if category != None else db[get_column_letter(5) + str(index)].value
            db[get_column_letter(6) + str(index)] = delivered if delivered != None else db[get_column_letter(6) + str(index)].value
            db[get_column_letter(9) + str(index)] = f"{date()} / {time()}"
            balance(value, ie, delivered)
        
    read()

    wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')
    
def delete():
    wb = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db = wb.active

    entry_list = []
    
    print("Available entries:")
    for column in db.iter_cols(min_row=2, max_row=db.max_row, min_col=2, max_col=2):
        for cell in column:
            if cell.value != None:
                entry_list.append(cell.value)

    print(entry_list)

    selected = str(input('Select entry to delete: ')).strip().upper()

    if selected not in entry_list:
        print('Entry does not exist.')
    else:
        index = entry_list.index(selected) + 2 
        db[get_column_letter(2) + str(index)].value = None
        db[get_column_letter(3) + str(index)].value = None
        db[get_column_letter(4) + str(index)].value = None
        db[get_column_letter(5) + str(index)].value = None
        db[get_column_letter(6) + str(index)].value = None
        db[get_column_letter(7) + str(index)].value = None
        db[get_column_letter(8) + str(index)].value = None
        balance(db[get_column_letter(5) + str(index)].value, 'E' if db[get_column_letter(3) + str(index)].value == 'I' else 'I', db[get_column_letter(7) + str(index)].value)
        
    read()
    wb.save('NSB-AlgoProg-Group-Work/Database.xlsx')

# TESTING
delete()

# NOTE:
# CODE HAS NOT BEEN FULLY TESTED YET // STILL NEED MORE TESTING
# ADDITION OF BALANCE FUNCTIONALITY IS REQUIRED