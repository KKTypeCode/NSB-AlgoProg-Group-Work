from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime as dt
import MainIO as io
async def pull_recurring(name=None, date_period=None, ie=None, value=None, category=None):
    # GET DATA FROM RECURRING DATABASE
    wb1 = load_workbook('NSB-AlgoProg-Group-Work/Recurring.xlsx')
    db1 = wb1.active
    
    entry_list = []
    date_list = []
    ie_list = []
    value_list = []
    category_list = []
    total = 0
    
    for columnname in db1.iter_cols(min_row=2, min_col=2, max_col=2):
    # GET THE ENTRY NAMES
        for cellname in columnname:
            if cellname.value != None:
                entry_list.append(cellname.value)
                total += 1
    
    for columndate in db1.iter_cols(min_row=2, max_row=db1.max_row, min_col=3, max_col=3):
    # GET THE ENTRY NAMES
        for celldate in columndate:
            if celldate.value != None:
                date_list.append(celldate.value)
                total += 1
    
    for columnie in db1.iter_cols(min_row=2, max_row=db1.max_row, min_col=4, max_col=4):
    # GET THE ENTRY NAMES
        for cellie in columnie:
            if cellie.value != None:
                ie_list.append(cellie.value)
                total += 1
                
    for columnvalue in db1.iter_cols(min_row=2, max_row=db1.max_row, min_col=5, max_col=5):
    # GET THE ENTRY NAMES
        for cellvalue in columnvalue:
            if cellvalue.value != None:
                value_list.append(cellvalue.value)
                total += 1
    
    for columncategory in db1.iter_cols(min_row=2, max_row=db1.max_row, min_col=6, max_col=6):
    # GET THE ENTRY NAMES
        for cellcategory in columncategory:
            if cellcategory.value != None:
                category_list.append(cellcategory.value)
                total += 1
    
    wb1.save('NSB-AlgoProg-Group-Work/Recurring.xlsx')
    wb1.close()
    
    # TRANSFER DATA TO MAIN DATABASE
    wb2 = load_workbook('NSB-AlgoProg-Group-Work/Database.xlsx')
    db2 = wb2.active
    
    total_entry = total // 5
    
    for each_entry in range(total_entry):
        if dt.datetime.now().date() == dt.datetime.strptime(date_list[each_entry], '%Y-%m-%d').date():
            db2[get_column_letter(2) + str(db2.max_row + 1)] = entry_list[each_entry]
            db2[get_column_letter(3) + str(db2.max_row + 1)] = date_list[each_entry]
            db2[get_column_letter(4) + str(db2.max_row + 1)] = ie_list[each_entry]
            db2[get_column_letter(5) + str(db2.max_row + 1)] = value_list[each_entry]
            db2[get_column_letter(6) + str(db2.max_row + 1)] = category_list[each_entry]
            db2[get_column_letter(7) + str(db2.max_row + 1)] = io.gen_id_key()
            db2[get_column_letter(8) + str(db2.max_row + 1)] = True
            db2[get_column_letter(9) + str(db2.max_row + 1)] = f"{io.date()} / {io.time()}"
            io.balance(value_list[each_entry], ie_list[each_entry], True)        
    
    wb2.save('NSB-AlgoProg-Group-Work/Database.xlsx')
    wb2.close()
