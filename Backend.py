from openpyxl import * #type: ignore
from openpyxl.utils import get_column_letter

"""
    REMINDER:
    - DO NOT FORGET TO INSTALL OPENPYXL IN YOUR MACHINE!
    - INSTALL IN CMD WITH: 'pip install openpyxl' / 'pip3 install openpyxl' /
            'python -m pip install openpyxl' / 'python3 -m pip install openpyxl'.
    - THE MODULE'S FILE MANIPULATOR CAN ONLY WORK WITH .xlsx FILE FORMATS (2010 ABOVE)
      AND CANNOT WORK WITH OTHER EXCEL FILE FORMATS.
    - CODE CANNOT WORK WHEN EXCEL FILE IS BEING OPENED.
"""

wb = load_workbook('Database.xlsx')
db = wb.active # TO MAKE FILE EDITABLE

# WORKSPACE

rowspan = int(input())
colspan = int(input())

for row in range(2, rowspan + 2): # START AT 2 BECAUSE INPUT START AT ROW 2
    for col in range(2, colspan + 2): # START AT 2 BECAUSE INPUT START AT COL B
        char = get_column_letter(col)
        

wb.save('Database.xlsx') # SAVE ALL CHANGES